#!/usr/bin/env python
__author__ = 'Timothy Lurvey'

import sys, os
import shutil
import openpyxl
import openpyxl.utils
import numpy as np

from _code.utils.utilities import configure_logger

ROSTER_SHEET = 'ROSTERS'
LEVEL = 'INFO'
log = configure_logger(name=__name__,
                       console_level=LEVEL,
                       log_file="{}.log".format(__file__))

base_path = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
file_path = os.path.join(base_path, 'mossixls', 'MILWAUKEE_FULL_STATS2.xlsx')

def build_roster_dict(ids: tuple, teams: tuple):
    """ {Player_ID: Team Name}"""
    if not len(ids) == len(teams):
        msg = f"ID and TEAMS are different sizes from {ROSTER_SHEET} sheet"
        log.error(msg)
        raise ValueError(msg)

    dplayers = {}
    for i in range(2, len(ids) - 1):
        if ids[i].value and teams[i].value:
            log.debug(f"found {ids[i].value:>9} for team {teams[i].value}")
            dplayers.update({ids[i].value: teams[i].value})
        else:
            log.debug(f"end of roster list.  {i} items")
            break

    log.info("Found {n:000d} players".format(n=len(dplayers)))
    return dplayers


def populate_players(wb: openpyxl.workbook, roster: dict, bat_or_pit: str = 'batting'):
    """Using a roster dictionary and the year, populate the roster for current and future years"""
    # get variables
    xyear = wb['DRAFT_STATS']['C3'].value
    log.info("Working on year {y} for {s}".format(y=xyear, s=bat_or_pit))
    sheet = wb['WAR_{}'.format(bat_or_pit)]
    col = openpyxl.utils.get_column_letter([x.value for x in sheet['A1:AZ1'][0]].index('ROSTER') + 1)
    # get arrays from worksheet column data
    plyr = np.array([x[0].value for x in sheet['A2:A{}'.format(sheet.max_row)]])
    year = np.array([x[0].value for x in sheet['E2:E{}'.format(sheet.max_row)]])
    # find matches
    match_name = np.in1d(plyr, np.array(list(roster.keys())))
    match_year = year >= xyear
    match_index = np.nonzero(match_name * match_year)[0]
    log.debug("Found {n} matches in spreadsheet".format(n=len(match_index)))
    # populate team names for matches
    for i, x in enumerate(match_year):
        if x:
            val = ""
            if i in match_index:
                val += roster.get(sheet['A{}'.format(i + 2)].value)
            sheet['{c}{r}'.format(c=col, r=i + 2)].value = val
            log.debug("set value for: {p} in {c}{r}".format(p=val, c=col, r=i + 2))

    # for i in match_index + 2:  # adjust for cells starting with 1
    #     sheet['{c}{r}'.format(c=col, r=i)].value = roster.get(sheet['A{}'.format(i)].value)


def main(excel_file=""):

    if not os.path.exists(excel_file):
        raise FileNotFoundError(f"file:{excel_file}")
    else:
        log.info(f"Working on {excel_file}")

    shutil.copy(src=excel_file, dst=excel_file + "_bak")
    log.info('opening workbook for reading...')
    r_wb = openpyxl.load_workbook(excel_file, data_only=True)
    log.info('begin read')
    #
    rost_dict = build_roster_dict(teams=r_wb['ROSTERS']['B:B'], ids=r_wb['ROSTERS']['C:C'])
    log.info('rosters built')
    r_wb.close()
    log.info('closing workbook for reading')
    #
    log.info('opening workbook for writing')
    w_wb = openpyxl.load_workbook(excel_file, data_only=False)
    log.info('begin write')
    populate_players(wb=w_wb, roster=rost_dict, bat_or_pit='batting')
    log.info('batters populated')
    #
    populate_players(wb=w_wb, roster=rost_dict, bat_or_pit='pitching')
    log.info('pitchers populated')
    #
    log.info('saving...')
    w_wb.save(excel_file)
    log.info('done!')


if __name__ == '__main__':
    main(excel_file=file_path)
